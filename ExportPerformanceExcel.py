#!/usr/bin/env python3
"""
Export performance comparison table to Excel.

Produces three sheets:
  1. Summary       – mean ± std per (controller, N) for all key metrics
  2. Relative      – TL vs BC percentage differences per N (TL as % of BC baseline)
  3. Per-Run Data  – raw values for every aligned individual run
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------- Paths & configuration -------------------------------------------

PROJECT_DIR = Path(__file__).resolve().parent
BASE_DIR = (
    PROJECT_DIR
    / "simulation_results"
    / "Final Results"
    / "Noise 0.1 22-05-2026"
)

BATCHES: dict[tuple[str, int], Path] = {
    ("Trafficlight",   37): BASE_DIR / "Trafficlight"   / "batch_20260521_164039_N37_T5400_R250p0",
    ("Trafficlight",   47): BASE_DIR / "Trafficlight"   / "batch_20260519_215956_N47_T5400_R250p0",
    ("Trafficlight",   57): BASE_DIR / "Trafficlight"   / "batch_20260520_203223_N57_T5400_R250p0",
    ("Barriercontrol", 37): BASE_DIR / "Barriercontrol" / "batch_20260521_094829_N37_T5400_R250p0",
    ("Barriercontrol", 47): BASE_DIR / "Barriercontrol" / "batch_20260519_155403_N47_T5400_R250p0",
    ("Barriercontrol", 57): BASE_DIR / "Barriercontrol" / "batch_20260520_151150_N57_T5400_R250p0",
}

OUTPUT_PATH = BASE_DIR / "performance_comparison.xlsx"

CTRL_DISPLAY = {"Trafficlight": "Traffic Light", "Barriercontrol": "Barrier Control"}

# Metrics from the batch summary CSV
METRICS = [
    ("final_throughput",    "Throughput [veh/s]",        "higher = better"),
    ("avg_speed_kmh",       "Avg Speed [km/h]",          "higher = better"),
    ("energy_kWh",          "Energy [kWh]",              "lower = better"),
    ("energy_per_km",       "Energy per km [kWh/km]",    "lower = better"),
    ("total_delay_s",       "Total Delay [s]",            "lower = better"),
    ("safety_violation_rate", "Safety Violation Rate",   "lower = better"),
    ("total_crossings",     "Total Crossings",            "higher = better"),
    ("total_distance_km",   "Total Distance [km]",        "higher = better"),
]

# For each metric: True if higher is better (used to colour cells)
HIGHER_BETTER = {m[0]: m[2] == "higher = better" for m in METRICS}

# ---------- Colour palette ---------------------------------------------------

CLR_HEADER_TL   = "2171B5"   # medium blue
CLR_HEADER_BC   = "CB181D"   # medium red
CLR_SECTION_HDR = "333333"
CLR_COL_HDR     = "555555"
CLR_BETTER      = "C6EFCE"   # light green fill
CLR_WORSE       = "FFC7CE"   # light red fill
CLR_NEUTRAL     = "FFEB9C"   # light yellow fill (near-zero difference)
CLR_WHITE       = "FFFFFF"
CLR_LIGHT_GREY  = "F2F2F2"

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------- Loading ----------------------------------------------------------

def load_summary(batch_dir: Path) -> pd.DataFrame:
    frames = []
    for sf in sorted(batch_dir.glob("batch_summary_*.csv")):
        try:
            df = pd.read_csv(sf)
            if not df.empty:
                frames.append(df)
        except Exception:
            pass
    if not frames:
        raise FileNotFoundError(f"No batch summary in {batch_dir}")
    summary = pd.concat(frames, ignore_index=True).sort_values("run_index")
    complete = summary[
        summary["final_time_seconds"] >= summary["duration_s"] * 0.99
    ].reset_index(drop=True)
    complete["energy_per_km"] = complete["energy_kWh"] / complete["total_distance_km"]
    return complete


def load_all() -> dict[tuple[str, int], pd.DataFrame]:
    data: dict[tuple[str, int], pd.DataFrame] = {}
    for key, path in BATCHES.items():
        ctrl, n = key
        print(f"Loading {CTRL_DISPLAY[ctrl]} N={n} …")
        try:
            data[key] = load_summary(path)
            print(f"  → {len(data[key])} complete runs")
        except FileNotFoundError as exc:
            print(f"  Skipping: {exc}")
    return data


def align_runs(
    data: dict[tuple[str, int], pd.DataFrame],
) -> dict[tuple[str, int], pd.DataFrame]:
    """Keep only run indices present in both controllers for each N."""
    n_values = {n for _, n in data}
    for n in n_values:
        tl_key = ("Trafficlight",   n)
        bc_key = ("Barriercontrol", n)
        if tl_key not in data or bc_key not in data:
            continue
        common = set(data[tl_key]["run_index"]) & set(data[bc_key]["run_index"])
        excluded_tl = set(data[tl_key]["run_index"]) - common
        excluded_bc = set(data[bc_key]["run_index"]) - common
        if excluded_tl:
            print(f"  Align: dropping TL N={n} runs {sorted(excluded_tl)}")
        if excluded_bc:
            print(f"  Align: dropping BC N={n} runs {sorted(excluded_bc)}")
        data[tl_key] = data[tl_key][data[tl_key]["run_index"].isin(common)].reset_index(drop=True)
        data[bc_key] = data[bc_key][data[bc_key]["run_index"].isin(common)].reset_index(drop=True)
    return data


# ---------- Excel helpers ----------------------------------------------------

def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col)


def _write(ws, row: int, col: int, value, bold=False, fill_hex=None,
           font_color="000000", align="left", num_fmt=None, border=True):
    c = _cell(ws, row, col)
    c.value = value
    c.font = Font(bold=bold, color=font_color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if fill_hex:
        c.fill = PatternFill("solid", fgColor=fill_hex)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = BORDER
    return c


def _autofit(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


def _pct_fill(pct: float, higher_better: bool) -> str:
    """Return fill colour based on direction of difference."""
    threshold = 1.0  # percent
    is_better = (pct > threshold) if higher_better else (pct < -threshold)
    is_worse  = (pct < -threshold) if higher_better else (pct > threshold)
    if is_better:
        return CLR_BETTER
    if is_worse:
        return CLR_WORSE
    return CLR_NEUTRAL


# ---------- Sheet 1: Summary ------------------------------------------------

def write_summary_sheet(ws, data: dict[tuple[str, int], pd.DataFrame]) -> None:
    ws.title = "Summary"
    ws.freeze_panes = "C3"

    metric_keys  = [m[0] for m in METRICS]
    metric_labels = [m[1] for m in METRICS]
    metric_notes  = [m[2] for m in METRICS]

    # Row 1 – title
    _write(ws, 1, 1, "Performance Summary — Barrier Control vs. Traffic Light",
           bold=True, fill_hex=CLR_SECTION_HDR, font_color=CLR_WHITE, align="center",
           border=False)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=2 + len(metric_keys) * 2)

    # Row 2 – column headers
    _write(ws, 2, 1, "Controller",      bold=True, fill_hex=CLR_COL_HDR, font_color=CLR_WHITE, align="center")
    _write(ws, 2, 2, "N  (# vehicles)", bold=True, fill_hex=CLR_COL_HDR, font_color=CLR_WHITE, align="center")
    _write(ws, 2, 3, "# Runs",          bold=True, fill_hex=CLR_COL_HDR, font_color=CLR_WHITE, align="center")

    col = 4
    for label, note in zip(metric_labels, metric_notes):
        _write(ws, 2, col,     f"Mean\n{label}", bold=True, fill_hex=CLR_COL_HDR,
               font_color=CLR_WHITE, align="center")
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center",
                                                         vertical="center", wrap_text=True)
        _write(ws, 2, col + 1, f"Std\n{label}",  bold=True, fill_hex=CLR_COL_HDR,
               font_color=CLR_WHITE, align="center")
        ws.cell(row=2, column=col + 1).alignment = Alignment(horizontal="center",
                                                              vertical="center", wrap_text=True)
        col += 2
    ws.row_dimensions[2].height = 36

    n_values = sorted({n for _, n in data})
    row = 3
    for n in n_values:
        for ctrl in ("Barriercontrol", "Trafficlight"):
            key = (ctrl, n)
            if key not in data:
                continue
            df = data[key]
            hdr_color = CLR_HEADER_BC if ctrl == "Barriercontrol" else CLR_HEADER_TL
            bg = CLR_LIGHT_GREY if row % 2 == 0 else CLR_WHITE
            _write(ws, row, 1, CTRL_DISPLAY[ctrl], bold=(ctrl == "Barriercontrol"),
                   fill_hex=hdr_color, font_color=CLR_WHITE, align="center")
            _write(ws, row, 2, n,          fill_hex=bg, align="center")
            _write(ws, row, 3, len(df),    fill_hex=bg, align="center")
            col = 4
            for mk in metric_keys:
                if mk in df.columns:
                    mean_val = df[mk].mean()
                    std_val  = df[mk].std()
                    fmt = "0.0000" if mk == "final_throughput" else (
                          "0.00"   if mk in ("energy_kWh", "energy_per_km",
                                              "total_distance_km", "avg_speed_kmh") else
                          "0.000"  if mk == "safety_violation_rate" else "0")
                    _write(ws, row, col,     mean_val, fill_hex=bg, align="right", num_fmt=fmt)
                    _write(ws, row, col + 1, std_val,  fill_hex=bg, align="right", num_fmt=fmt)
                else:
                    _write(ws, row, col,     "N/A", fill_hex=bg)
                    _write(ws, row, col + 1, "N/A", fill_hex=bg)
                col += 2
            row += 1

        # Spacer row between N groups
        for c in range(1, col):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="EEEEEE")
        row += 1

    # Footer: metric notes
    row += 1
    _write(ws, row, 1, "Metric notes:", bold=True, border=False)
    for i, (mk, label, note) in enumerate(METRICS):
        _write(ws, row + 1 + i, 1, label,  border=False)
        _write(ws, row + 1 + i, 2, f"({note})", border=False)

    _autofit(ws)


# ---------- Sheet 2: Relative performance ------------------------------------

def write_relative_sheet(ws, data: dict[tuple[str, int], pd.DataFrame]) -> None:
    ws.title = "Relative Performance"
    ws.freeze_panes = "B3"

    metric_keys   = [m[0] for m in METRICS]
    metric_labels = [m[1] for m in METRICS]

    # Title
    n_cols = 1 + len(metric_keys) * 3 + 1
    _write(ws, 1, 1,
           "Relative Performance: Traffic Light vs. Barrier Control  "
           "(positive % = TL is higher; negative % = TL is lower)",
           bold=True, fill_hex=CLR_SECTION_HDR, font_color=CLR_WHITE,
           align="center", border=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Column headers
    _write(ws, 2, 1, "N", bold=True, fill_hex=CLR_COL_HDR, font_color=CLR_WHITE, align="center")
    col = 2
    for label in metric_labels:
        _write(ws, 2, col,     f"BC mean\n{label}",     bold=True, fill_hex=CLR_HEADER_BC,
               font_color=CLR_WHITE, align="center")
        _write(ws, 2, col + 1, f"TL mean\n{label}",     bold=True, fill_hex=CLR_HEADER_TL,
               font_color=CLR_WHITE, align="center")
        _write(ws, 2, col + 2, f"TL vs BC\n% diff",     bold=True, fill_hex=CLR_COL_HDR,
               font_color=CLR_WHITE, align="center")
        for c in (col, col + 1, col + 2):
            ws.cell(row=2, column=c).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
        col += 3
    ws.row_dimensions[2].height = 36

    n_values = sorted({n for _, n in data})
    row = 3
    for n in n_values:
        bc_key = ("Barriercontrol", n)
        tl_key = ("Trafficlight",   n)
        if bc_key not in data or tl_key not in data:
            continue
        bc_df = data[bc_key]
        tl_df = data[tl_key]

        bg = CLR_LIGHT_GREY if row % 2 == 0 else CLR_WHITE
        _write(ws, row, 1, n, bold=True, fill_hex=bg, align="center")
        col = 2
        for mk in metric_keys:
            higher_better = HIGHER_BETTER[mk]
            fmt_val = ("0.0000" if mk == "final_throughput" else
                       "0.00"   if mk in ("energy_kWh", "energy_per_km",
                                           "total_distance_km", "avg_speed_kmh") else
                       "0.000"  if mk == "safety_violation_rate" else "0")
            fmt_pct = '+0.0%;-0.0%;"—"'

            if mk in bc_df.columns and mk in tl_df.columns:
                bc_mean = bc_df[mk].mean()
                tl_mean = tl_df[mk].mean()
                if bc_mean != 0:
                    pct = (tl_mean - bc_mean) / abs(bc_mean)
                else:
                    pct = float("nan")
                _write(ws, row, col,     bc_mean, fill_hex=bg,  align="right", num_fmt=fmt_val)
                _write(ws, row, col + 1, tl_mean, fill_hex=bg,  align="right", num_fmt=fmt_val)
                pct_fill = _pct_fill(pct * 100, higher_better) if not np.isnan(pct) else CLR_NEUTRAL
                _write(ws, row, col + 2, pct,     fill_hex=pct_fill, align="center",
                       bold=True, num_fmt=fmt_pct)
            else:
                for c in (col, col + 1, col + 2):
                    _write(ws, row, c, "N/A", fill_hex=bg)
            col += 3
        row += 1

    # Legend
    row += 1
    _write(ws, row, 1, "Colour legend:", bold=True, border=False)
    for fill, desc in [
        (CLR_BETTER,  "Green  → Traffic Light performs better than Barrier Control"),
        (CLR_WORSE,   "Red    → Traffic Light performs worse than Barrier Control"),
        (CLR_NEUTRAL, "Yellow → Difference < 1 % (negligible)"),
    ]:
        row += 1
        c = ws.cell(row=row, column=1)
        c.fill = PatternFill("solid", fgColor=fill)
        c.value = desc
        c.border = BORDER

    row += 2
    _write(ws, row, 1,
           "Note: For 'lower = better' metrics (energy, delay), a negative % means TL uses less → better.",
           border=False)

    _autofit(ws)


# ---------- Sheet 3: Per-run data -------------------------------------------

def write_per_run_sheet(ws, data: dict[tuple[str, int], pd.DataFrame]) -> None:
    ws.title = "Per-Run Data"
    ws.freeze_panes = "E3"

    metric_keys   = [m[0] for m in METRICS]
    metric_labels = [m[1] for m in METRICS]

    # Title
    _write(ws, 1, 1, "Per-Run Raw Data — all aligned runs",
           bold=True, fill_hex=CLR_SECTION_HDR, font_color=CLR_WHITE,
           align="center", border=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=4 + len(metric_keys))

    headers = ["Controller", "N (# vehicles)", "Run Index", "# Vehicles (actual)"] + metric_labels
    for col, hdr in enumerate(headers, start=1):
        _write(ws, 2, col, hdr, bold=True, fill_hex=CLR_COL_HDR,
               font_color=CLR_WHITE, align="center")
        ws.cell(row=2, column=col).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    n_values = sorted({n for _, n in data})
    row = 3
    for n in n_values:
        for ctrl in ("Barriercontrol", "Trafficlight"):
            key = (ctrl, n)
            if key not in data:
                continue
            df = data[key]
            hdr_color = CLR_HEADER_BC if ctrl == "Barriercontrol" else CLR_HEADER_TL
            for _, run_row in df.iterrows():
                bg = CLR_LIGHT_GREY if row % 2 == 0 else CLR_WHITE
                _write(ws, row, 1, CTRL_DISPLAY[ctrl], fill_hex=hdr_color,
                       font_color=CLR_WHITE, align="center")
                _write(ws, row, 2, n,                       fill_hex=bg, align="center")
                _write(ws, row, 3, int(run_row["run_index"]), fill_hex=bg, align="center")
                _write(ws, row, 4, int(run_row.get("vehicles", n)), fill_hex=bg, align="center")
                for col_offset, mk in enumerate(metric_keys, start=5):
                    val = run_row.get(mk, None)
                    fmt = ("0.0000" if mk == "final_throughput" else
                           "0.00"   if mk in ("energy_kWh", "energy_per_km",
                                               "total_distance_km", "avg_speed_kmh") else
                           "0.000"  if mk == "safety_violation_rate" else "0")
                    _write(ws, row, col_offset, val if pd.notna(val) else "N/A",
                           fill_hex=bg, align="right",
                           num_fmt=fmt if pd.notna(val) else None)
                row += 1
        # Spacer between N groups
        row += 1

    _autofit(ws)


# ---------- Main ------------------------------------------------------------

def main() -> None:
    print("=== Loading batch summaries ===")
    data = load_all()

    print("\n=== Aligning run indices across controllers ===")
    data = align_runs(data)

    print("\n=== Run counts after alignment ===")
    for n in sorted({n for _, n in data}):
        for ctrl in ("Barriercontrol", "Trafficlight"):
            key = (ctrl, n)
            if key in data:
                runs = sorted(data[key]["run_index"].tolist())
                print(f"  {CTRL_DISPLAY[ctrl]} N={n}: {len(runs)} runs  {runs}")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    print("\n=== Writing Excel ===")
    write_summary_sheet(wb.create_sheet(), data)
    print("  ✓ Summary sheet")
    write_relative_sheet(wb.create_sheet(), data)
    print("  ✓ Relative Performance sheet")
    write_per_run_sheet(wb.create_sheet(), data)
    print("  ✓ Per-Run Data sheet")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nSaved → {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
